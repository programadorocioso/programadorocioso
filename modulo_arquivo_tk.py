from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
from openpyxl import Workbook, load_workbook
import os
def leitura_escrita_excell(linhas):
    if os.path.exists("cotacao.xlsx"):
        planilha = load_workbook(filename="cotacao.xlsx")
        celula = planilha.active
        celula.column_dimensions['A'].widht = 20
        celula.column_dimensions['B'].widht = 20
        celula.column_dimensions['C'].widht = 60
        for linha in linhas:
            planilha.active.append(linha)
    else:
        planilha = Workbook()
        celula = planilha.active
        celula.column_dimensions['A'].widht = 20
        celula.column_dimensions['B'].widht = 20
        celula.column_dimensions['C'].widht = 60
        celula.append(["Ativo","Valor","Data e hora da Consulta"])
        for linha in linhas:
            celula = planilha.active
            celula.append(linha)
    planilha.save("cotacao.xlsx")
def ler_para_lista():
    planilha = load_workbook("cotacao.xlsx")
    celula_ativa = planilha.active
    linhas = []
    for row in celula_ativa.iter_rows():
        linha = []
        for celula in row:
            linha.append(celula.value)
        linhas.append(linha)
    return linhas
def xlsx_para_pdf():
    if os.path.exists("cotacao.xlsx"):
        dados = ler_para_lista()
        documento = SimpleDocTemplate("cotacao.pdf", pagesize=letter)
        tabela = Table(dados)
        documento.build([tabela])
def arquivo_texto(local):
    texto = open(local,"r").readlines()
    return texto
def lista_string(lista):
    palavras = []
    for linha in lista:
        if linha.__contains__(" "):
            temp = linha.replace("\n","").split()
            for i in temp:
                palavras.append(i)
        else:
            palavras.append(linha)
    return palavras
def string_lista(lista):
    linha = ""
    for i in lista:
        if i.__contains__("\n"):
            i.replace("\n","")
        linha = linha+str(i)+" "
        if i == " ":
            continue
        if i == "":
            continue
    return linha